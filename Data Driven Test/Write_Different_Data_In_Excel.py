# Install openpyxl module to work with excel

import openpyxl

file = "C:\\Users\\ialmu\\PycharmProjects\\mutipledata.xlsx"

#file -> Workbook -> Sheets -> rows -> Cells
workbook = openpyxl.load_workbook(file)  #will load the workbook from the file
sheet = workbook["Sheet1"]      #we can use sheet name instead of sheet1 but sheet1 is recommended in case of a lot of sheet
#sheet.workbook.active #alternative way of sheet = workbook["Sheet1"]

sheet.cell(1,1).value =123                         #row and clomun
sheet.cell(1,2).value ="Imran"
sheet.cell(1,3).value ="Engineer"

sheet.cell(2,1).value =567
sheet.cell(2,2).value ="Sajib"
sheet.cell(2,3).value ="test engineer"

sheet.cell(3,1).value =891
sheet.cell(3,2).value ="Munyeem"
sheet.cell(3,3).value ="Devops"

sheet.cell(1,1).value =123
sheet.cell(1,1).value =123

workbook.save(file)             #save data in excel