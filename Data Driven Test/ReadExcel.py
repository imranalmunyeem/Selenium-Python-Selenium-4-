# Install openpyxl module to work with excel

import openpyxl

file = "C:\\Users\\ialmu\\PycharmProjects\\data.xlsx"

#file -> Workbook -> Sheets -> rows -> Cells
workbook = openpyxl.load_workbook(file)  #will load the workbook from the file
sheet = workbook["Sheet1"]

#count number of rows and column in the sheet
rows = sheet.max_row
columns = sheet.max_column

#read all the rows and column from excel sheet
for row in range (1,rows+1):   #start from row1. and we used rows+1 because it will count n-1
    for column in range (1,columns+1):
        print(sheet.cell(row,column).value) #it will take the value of row and clomuns from each cells of the sheet
       #print(sheet.cell(row,column).value,end='        ') #to show the result in table format
        print()                             #to keep an space between each line


